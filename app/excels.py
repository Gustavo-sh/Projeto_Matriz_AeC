import os
import openpyxl
from openpyxl import Workbook

USERS_FILE = "usuarios.xlsx"
REGISTROS_FILE = "registros.xlsx"

def generate_users_excel():
    """Garante que o arquivo de usuários exista com cabeçalho"""
    if not os.path.exists(USERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["username", "password", "role"])
        wb.save(USERS_FILE)

def save_user(username, password, role=""):
    """Salva usuário no Excel (aceita role opcional para compatibilidade)"""
    if not os.path.exists(USERS_FILE):
        generate_users_excel()

    wb = openpyxl.load_workbook(USERS_FILE)
    ws = wb.active
    ws.append([username, password, role])
    wb.save(USERS_FILE)
    wb.close()


def get_user(username):
    """Recupera usuário pelo nome, retornando também a role se existir"""
    if not os.path.exists(USERS_FILE):
        return None

    wb = openpyxl.load_workbook(USERS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            password = row[1] if len(row) > 1 else None
            role = row[2] if len(row) > 2 else ""
            wb.close()
            return {"username": row[0], "password": password, "role": role}
    wb.close()
    return None


def generate_registros_excel():
    """Garante que o arquivo de registros exista com cabeçalho"""
    if not os.path.exists(REGISTROS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"
        headers = [
            "Id Nome Indicador", "Meta", "Moeda", "Critério Final", "Área",
            "Tipo Faturamento", "Classificação", "Meta Classificação",
            "Escala", "Acumulado", "Tipo Indicador","Data Início", "Data Fim", "Tipo Matriz",
            "Esquema Acumulado", "Descrição", "Ativo", "Chamado",
            "Possui DMM", "Multiplicador", "DMM", "Atributo", "Matricula"
        ]
        ws.append(headers)
        wb.save(REGISTROS_FILE)

def save_registros_to_excel(registros, matricula):
    """Salva registros no Excel, sempre appending no final, garantindo fechamento do arquivo"""
    if not os.path.exists(REGISTROS_FILE):
        raise FileNotFoundError(
            "O arquivo registros.xlsx não existe. Rode generate_registros_excel primeiro."
        )

    wb = None
    try:
        wb = openpyxl.load_workbook(REGISTROS_FILE)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        mapping = {
            "Id Nome Indicador": "nome",
            "Meta": "meta",
            "Moeda": "moeda",
            "Critério Final": "criterio_final",
            "Área": "area",
            "Tipo Faturamento": "tipo_faturamento",
            "Classificação": "classificacao",
            "Meta Classificação": "meta_classificacao",
            "Escala": "escala",
            "Acumulado": "acumulado",
            "Data Início": "data_inicio",
            "Data Fim": "data_fim",
            "Tipo Matriz": "tipo_matriz",
            "Esquema Acumulado": "esquema_acumulado",
            "Descrição": "descricao",
            "Ativo": "ativo",
            "Chamado": "chamado",
            "Possui DMM": "possuiDmm",
            "Multiplicador": "mult",
            "DMM": "dmm",
            "Atributo": "atributo",
            "Matricula": "matricula",
            "Tipo Indicador": "tipo_indicador"
        }

        for r in registros:
            row_data = []
            for col in headers:
                if col == "Matricula":
                    row_data.append(matricula)
                else:
                    key = mapping.get(col)
                    row_data.append(r.get(key, "") if key else "")
            ws.append(row_data)

        wb.save(REGISTROS_FILE)
        return REGISTROS_FILE

    finally:
        if wb:
            wb.close()
